import openpyxl
import pandas as pd


arquivo_excel = 'danos_TESTES.xlsx'

def get_data() -> pd.DataFrame:
    with pd.ExcelFile('danos_TESTES.xlsx') as xls:
        return pd.read_excel(xls)

def get_parcela():
    df = get_data()
    
    return list(df.PARCELA.dropna().unique())

def get_planta():
    df = get_data()
    
    return list(df.PLANTA.dropna().unique())

def get_dano():
    df = get_data()
    
    return list(df.NOTA_DANO.dropna().unique())

def get_datas():
    df = get_data()

    return list(df.DATA.dropna().unique())

def adicionar_linha_excel(dados):
    workbook = openpyxl.load_workbook(arquivo_excel)
    planilha = workbook.active
    ultima_linha = planilha.max_row + 1

    for index, item in enumerate(dados):
        planilha.cell(row=ultima_linha, column=index+1, value=item)
    

    workbook.save(arquivo_excel)


def deletar_linha_excel(data, parcela, planta):
    workbook = openpyxl.load_workbook(arquivo_excel)
    planilha = workbook.active
    
    linha_a_deletar = None
    for row in range(2, planilha.max_row + 1):
        if (
            planilha.cell(row=row, column=1).value == data and
            planilha.cell(row=row, column=2).value == parcela and
            planilha.cell(row=row, column=3).value == planta
        ):
            linha_a_deletar = row
            break

    if linha_a_deletar is not None:
        planilha.delete_rows(linha_a_deletar, 1)

        workbook.save(arquivo_excel)
        return True
    return False


def editar_linha_excel(dados_antigo, novos):
    workbook = openpyxl.load_workbook(arquivo_excel)
    planilha = workbook.active
    
    linha_a_editar = None
    for row in range(2, planilha.max_row + 1):
        if (
            planilha.cell(row=row, column=1).value == dados_antigo[0] and
            planilha.cell(row=row, column=2).value == dados_antigo[1] and
            planilha.cell(row=row, column=3).value == dados_antigo[2]
        ):
            for colum, novo in novos:
                planilha.cell(row=row, column=colum, value=novo)
            workbook.save(arquivo_excel)
            return True
    return False
    